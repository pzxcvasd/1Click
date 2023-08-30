import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Color, Font, Alignment
import re

framerate = 24

def timecode_to_frames(timecode):
    return sum(f * int(t) for f,t in zip((3600*framerate, 60*framerate, framerate, 1), timecode.split(':')))

def frames_to_TC(frames):
    h = int(frames / 86400)
    m = int(frames / 1440) % 60
    s = int((frames % 1440)/24)
    f = frames % 1440 % 24
    return ("%02d:%02d:%02d:%02d" % (h, m, s, f))

def listToString(str_list):
    result = ""
    for s in str_list:
            result += s + " "
    return result.strip()


wb = Workbook()
ws = wb.active

col_names = ['', 'Reel ID', 'TC in', 'TC out', 'Duration', 'frame']
col_width = [6, 40, 15,15,15,15]

for seq, name in enumerate(col_names):
        ws.cell(row=1, column=seq+1, value=name)
        ws.column_dimensions[get_column_letter(seq+1)].width = col_width[seq]
        ws.cell(row=1, column=seq+1).font = Font(bold=True)
        ws.cell(row=1, column=seq+1).fill = PatternFill(patternType='solid',fgColor=Color(indexed=31))

f = open('/Users/yeonsuseo/Desktop/CHS.edl')

lines = f.read().splitlines()
cnt = 1

for num, line in enumerate(lines) :

    line = line.split()
    
    if 'V' in line and 'C' in line :
        ws.cell(row = cnt+1, column = 1, value = cnt)
        ws. cell(row = cnt+1, column =2, value = listToString(re.findall('[A-Z]\w+',line[1])))
        ws.cell(row = cnt+1, column = 3, value = line[4])
        ws.cell(row = cnt+1, column = 4, value = line[5])
        start_frame = timecode_to_frames(line[4])
        end_frame = timecode_to_frames(line[5])
        duration_frame = end_frame - start_frame + 1
        duration_TC = frames_to_TC(duration_frame)
        ws.cell(row = cnt+1, column =5, value = duration_TC)
        ws.cell(row = cnt+1, column =6, value = duration_frame)

        cnt += 1
        


for row in ws.rows:
            for cell in row:
                if cell.value == 'Dissolve' :
                    cell.font = Font(color = "FF0000")
                cell.alignment = Alignment(
                    horizontal='center', vertical='center', wrap_text=True)

output_file = os.path.join(os.path.expanduser('~'),'Desktop',os.path.splitext(os.path.basename(f))[0])

wb.save(output_file)