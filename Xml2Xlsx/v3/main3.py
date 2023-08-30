from PySide6.QtWidgets import *
from ui_0929 import Ui_MainWindow
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Color, Font, Alignment
import xml.etree.ElementTree as qwe
import os
import math
import re


#======== 버전 정보 =========
# version 5   --2022.11.10
# Reel ID 제대로 찾아줌 (not 'Slug')
# Edit Duration. ORG Duration 위치 바꿈

#=================================================
def frames_to_TC(frames):
    h = int(frames / 86400)
    m = int(frames / 1440) % 60
    s = int((frames % 1440)/24)
    f = frames % 1440 % 24
    return ("%02d:%02d:%02d:%02d" % (h, m, s, f))
#=================================================


class myWindow(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super(myWindow, self).__init__()
        self.setupUi(self)
        self.Done.hide()
        self.xml_upload.clicked.connect(self.upload_xml)
        self.start.clicked.connect(self.toXlsx)

    def upload_xml(self):
        self.Done.hide()
        self.xmlPath.clear()
        global xmlfile
        xmlfile = QFileDialog.getOpenFileName(self, self.tr('Open File'), '/', 'All File(*);; xml File(*.xml)')[0]

        self.xmlPath.setText(xmlfile.split("/")[-1])

    def toXlsx(self):

        wb = Workbook()
        ws = wb.active

# ============================================SET UP======================================================

        tree = qwe.parse(xmlfile)
        K = tree.find("sequence/media/video/track")
        A = K.findall('clipitem')

        file_name_li = []
        issue_li = []

        wb = Workbook()
        ws = wb.active 
                  #   1      2          3           4        5            6(n)                  7(n)              8(6)           9(7)         10(8)           11(9)          12(10)         13(11)      14(12)
        col_names = ['', 'Reel ID', 'Clip name', 'TC in', 'TC out', 'Timeline TC in', 'Timeline TC out' ,'Duration', 'Duration(f)','ORG_Duration', 'ORG_Duration(f)', 'Retime (%)', 'Resize (%)', 'note']
        col_width = [6,     40,        40,          15,       15,        18,                  18,                 16,         17,             15, 16, 12, 12, 17]

        for seq, name in enumerate(col_names):
            ws.cell(row=1, column=seq+1, value=name)
            ws.column_dimensions[get_column_letter(seq+1)].width = col_width[seq]
            ws.cell(row=1, column=seq+1).font = Font(bold=True)
            ws.cell(row=1, column=seq+1).fill = PatternFill(patternType='solid',fgColor=Color(indexed=31))

# ============================================SET FIN======================================================
        
        for num, clip in enumerate(A) :
            issue_li.clear()

            global TC_in
            global TC_out

            try :
                ws.cell(row=num+2, column=2, value=''.join(re.findall('[A-Z]\w+',clip.find('file').find('timecode').find('reel').findtext('name'))))
                file_name_li.append([clip.find('file').get('id'),
                    ''.join(re.findall('[A-Z]\w+',clip.find('file').find('timecode').find('reel').findtext('name'))),
                    clip.find('file').find('timecode').findtext('frame') if clip.find('file').find('timecode').findtext('frame') != None else 0 ])
                TC_in = frames_to_TC(int(clip.find('file').find('timecode').findtext('frame')) + int(clip.findtext('in'))) if clip.find('file').find('timecode').findtext('frame') != None else frames_to_TC(int(clip.findtext('in')))
                TC_out = frames_to_TC(int(clip.find('file').find('timecode').findtext('frame')) + int(clip.findtext('out'))) if clip.find('file').find('timecode').findtext('frame') != None else frames_to_TC(int(clip.findtext('out')))

                ws.cell(row = num+2, column= 4, value=TC_in)
                ws.cell(row=num+2, column=5, value=TC_out)

            except :
                if clip.find('file').findtext('name') is None :
                    for i in range(len(file_name_li)) :
                        if clip.find('file').get('id') == file_name_li[i][0] :
                            ws.cell(row=num+2, column = 2, value = file_name_li[i][1])
                            TC_in = frames_to_TC(int(file_name_li[i][2]) + int(clip.findtext('in')))
                            TC_out = frames_to_TC(int(file_name_li[i][2]) + int(clip.findtext('out')))
                            TC_duration = int(clip.findtext('end')) - int(clip.findtext('start'))

                            ws.cell(row = num+2, column= 4, value=TC_in)
                            ws.cell(row=num+2, column=5, value=TC_out)


            ws.cell(row=num+2, column=1, value=num+1) #Event
            ws.cell(row=num+2, column=3, value=clip.findtext('name')) #Clip Name
            ws.cell(row=num+2, column=6, value = frames_to_TC(int(clip.findtext('start'))))
            ws.cell(row=num+2, column=7, value = frames_to_TC(int(clip.findtext('end'))-1))

            TC_duration = int(clip.findtext('end')) - int(clip.findtext('start'))
            ws.cell(row = num+2, column=9, value = TC_duration if int(clip.findtext('start')) >= 0 and int(clip.findtext('end')) >= 0 else "Dissolve") #Duration frame
            ws.cell(row = num+2, column=8, value = frames_to_TC(TC_duration) if int(clip.findtext('start')) >= 0 and int(clip.findtext('end')) >=0 else "Dissolve")

            filter_li=clip.findall('filter')

            if len(filter_li) > 0 :
                for filter in filter_li :
                    effect_li = filter.find('effect').findall('parameter')

                    for effect in effect_li :
                        if effect.findtext('name') == 'Scale' and filter.find('effect').findtext('effectid') != 'GraphicAndType':
                            ws.cell(row=num+2, column=13, value=float(effect.findtext('value'))) # Resize
                            if effect.find('keyframe') is not None :
                                issue_li.append('* scale variable')
                        if effect.findtext('name') == 'speed' :
                            ws.cell(row=num+2, column=12, value=float(effect.findtext('value'))) #Retime
                            retime = float(effect.findtext('value'))
                
                        if effect.findtext('name') == 'reverse' :
                            if effect.findtext('value') == 'TRUE' :
                                issue_li.append('* reverse')
                        if effect.findtext('name') == 'graphdict' :
                            speed_key = effect.findall('keyframe')
                            for keyframe in speed_key :
                                new_du = ws.cell(row=num+2, column=9).value
                                if keyframe.findtext('when') == clip.findtext('in') : #TC_in 다시
                                    edit_du_in = float(keyframe.findtext('value'))
                                    if clip.find('file').findtext('name') != None :
                                        new_TC_in = int(clip.find('file').find('timecode').findtext('frame') if clip.find('file').find('timecode').findtext('frame') != None else 0) + edit_du_in
                                    else :
                                        for i in range(len(file_name_li)) :
                                            if clip.find('file').get('id') == file_name_li[i][0] :
                                                new_TC_in = int(file_name_li[i][2]) + edit_du_in

                            if new_du != "Dissolve" :
                                if retime < 100 :
                                    new_frame = math.ceil((retime/100) * new_du)
                                elif retime > 100 :
                                    new_frame = math.floor(retime/100 * new_du)
                                ws.cell(row = num + 2, column=10, value= frames_to_TC(new_frame))
                                ws.cell(row = num + 2, column=11, value= new_frame)
                                ws.cell(row=num+2, column=4, value=frames_to_TC(new_TC_in))
                                ws.cell(row=num+2, column=5, value=frames_to_TC(new_TC_in + int(new_frame)))
            issue = '\n'.join(s for s in issue_li)
            ws.cell(row=num+2, column=14, value=issue)

        for row in ws.rows:
            for cell in row:
                if cell.value == 'Dissolve' :
                    cell.font = Font(color = "FF0000")
                cell.alignment = Alignment(
                    horizontal='center', vertical='center', wrap_text=True)

        user_name = os.path.expanduser('~') + '/Desktop/'
        file_name = 'xml_list.xlsx'
        output = user_name + file_name
        uniq = 1

        while os.path.exists(output) :
            file_name = 'xml_list(%d).xlsx' % uniq
            output = user_name + file_name
            uniq += 1

        wb.save(output)
        self.Done.setText('* 바탕화면에 %s 생성 완료.' %file_name)
        self.Done.show()



app = QApplication()
window = myWindow()
window.show()
app.exec()