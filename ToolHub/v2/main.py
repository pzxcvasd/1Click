from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Color, Font, Alignment
import xml.etree.ElementTree as qwe
from PySide6.QtWidgets import *
import re,os,math
from ui_ToolArchives_3 import Ui_MainWindow
import OpenEXR
import subprocess
import sys
import cv2

framerate = 24

def print_resolution(folder_path, row):
        exr_files = [file for file in os.listdir(folder_path) if file.endswith(".exr")]

        exr_path = os.path.join(folder_path, exr_files[0])
    
        exr_file = OpenEXR.InputFile(exr_path)
        dw = exr_file.header()['dataWindow']
        width = dw.max.x - dw.min.x + 1
        height = dw.max.y - dw.min.y + 1
    
        folder_name = os.path.basename(folder_path)
        ws.cell(row=row, column=1, value=folder_name)
        ws.cell(row=row, column=2, value=f"{width} x {height}")
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        row += 1
        # return row 

def timecode_to_frames(timecode):
    return sum(f * int(t) for f,t in zip((3600*framerate, 60*framerate, framerate, 1), timecode.split(':')))

def frames_to_TC(frames):
    h = int(frames / 86400)
    m = int(frames / 1440) % 60
    s = int((frames % 1440)/24)
    f = frames % 1440 % 24
    return ("%02d:%02d:%02d:%02d" % (h, m, s, f))

def listToString(str_list):
    result = ""
    for s in str_list:
            result += s + " "
    return result.strip()



class myWindow(QMainWindow, Ui_MainWindow):

    def check_wifi_status(self):
        target_wifi = "SPMCd"
        cmd = "networksetup -getairportnetwork en0"
        output = subprocess.check_output(cmd, shell=True).decode("utf-8")

        if target_wifi not in output:
            QMessageBox.critical(self, "Error", "SPMC 내부에서만 실행가능합니다.")
            sys.exit(1)

    def __init__(self):
        
        super(myWindow, self).__init__()
        self.setupUi(self)
        self.Tab.setCurrentIndex(0)

        # 모든 탭의 Done 버튼 숨기기
        self.Done_BI.hide()
        self.Done_RN.hide()
        self.Done_e2x.hide()
        self.Done_x2x.hide()
        self.Done_size.hide()
        self.Done_thumb.hide()
        self.label_thumb.hide()
        self.label_movPath.hide()
        # self.Done.hide() # plate check

        # 업로드 될 때 뜨는 파일명 숨기기
        self.list_label_BI.hide()
        self.list_label_RN.hide()
        self.folder_label_RN.hide()
        self.xml_label_BI.hide()
        self.thumb_title_2.hide()

        # 업로드 버튼 시그널 연결
        self.upload_e2x.clicked.connect(self.E2X_uploadEDL)
        self.upload_x2x.clicked.connect(self.X2X_uploadXML)
        # self.upload_pc.clicked.connect(self.PC_uploadXL)
        self.upload_pc_2.clicked.connect(self.Size_uploadEXR)

        self.list_BI.clicked.connect(self.BI_uploadList)
        self.xml_BI.clicked.connect(self.BI_uploadXML)

        self.list_RN.clicked.connect(self.RN_uploadList)
        self.folder_RN_2.clicked.connect(self.RN_uploadFolder)
        self.movPath.clicked.connect(self.Th_input)
        self.thumbPath.clicked.connect(self.Th_output)

      	# # 시작버튼 시그널 연결
        self.start_BI.clicked.connect(self.BI_start)
        self.start_e2x.clicked.connect(self.E2X_start)
        self.start_x2x.clicked.connect(self.X2X_start)
        self.start_RN.clicked.connect(self.RN_start)
        self.start_size.clicked.connect(self.Size_start)
        self.start_thumb.clicked.connect(self.Th_start)
        # self.start_pc.clicked.connect(self.PC_start)





    def Th_input(self) :
        self.label_thumb.hide()
        global movfolder
        movfolder = QFileDialog.getExistingDirectory(self, self.tr("Open Data files"), "/", QFileDialog.ShowDirsOnly)
        self.label_movPath.setText(movfolder.split("/")[-1])
        self.label_movPath.show()


    def Th_output(self) :
        self.label_thumb.hide()
        global outfolder
        outfolder = QFileDialog.getExistingDirectory(self, self.tr("Open Data files"), "/", QFileDialog.ShowDirsOnly)
        self.label_thumb.setText(outfolder.split("/")[-1])
        self.label_thumb.show()

    def Th_start(self) :
        def extract_thumbnail(input_folder, output_folder):
    # 폴더 구조를 그대로 복제
            os.makedirs(output_folder, exist_ok=True)

            for root, _, files in os.walk(input_folder):
                for file in files:
                    if file.endswith(".mov"):
                        file_path = os.path.join(root, file)

                # 상위 폴더의 이름을 추출하여 필요한 경우 폴더 생성
                        relative_path = os.path.relpath(root, input_folder)
                        output_subfolder = os.path.join(output_folder, relative_path)
                        os.makedirs(output_subfolder, exist_ok=True)

                # 썸네일 추출
                        video = cv2.VideoCapture(file_path)
                        success, image = video.read()
                        if success:
                    # 썸네일 파일 경로
                            thumbnail_path = os.path.join(output_subfolder, file[:-4] + ".jpg")
                            cv2.imwrite(thumbnail_path, image)    
        extract_thumbnail(movfolder,outfolder)
        self.Done_thumb.show()





   # ========== edl 2 xlsx ============

    def E2X_uploadEDL(self):
        self.Done_e2x.hide()
        self.edlpath.clear()
        global edlfile
        edlfile = QFileDialog.getOpenFileName(self, self.tr('Open File'), '/', 'All File(*);; edl File(*.edl)')[0]
        self.edlpath.setText(edlfile.split("/")[-1])

    def E2X_start(self):
        wb = Workbook()
        ws = wb.active

        col_names = ['', 'Reel ID', 'TC in', 'TC out', 'Duration', 'frame']
        col_width = [6, 40, 15,15,15,15]

        for seq, name in enumerate(col_names):
                ws.cell(row=1, column=seq+1, value=name)
                ws.column_dimensions[get_column_letter(seq+1)].width = col_width[seq]
                ws.cell(row=1, column=seq+1).font = Font(bold=True)
                ws.cell(row=1, column=seq+1).fill = PatternFill(patternType='solid',fgColor=Color(indexed=31))

        f = open(edlfile)

        lines = f.read().splitlines()
        cnt = 1

        for line in lines :

            line = line.split()
    
            if 'V' in line and 'C' in line :
                ws.cell(row = cnt+1, column = 1, value = cnt)
                ws. cell(row = cnt+1, column =2, value = listToString(re.findall('[A-Z]\w+',line[1])))
                ws.cell(row = cnt+1, column = 3, value = line[4])
                ws.cell(row = cnt+1, column = 4, value = line[5])
                start_frame = timecode_to_frames(line[4])
                end_frame = timecode_to_frames(line[5])
                duration_frame = end_frame - start_frame
                duration_TC = frames_to_TC(duration_frame)
                ws.cell(row = cnt+1, column =5, value = duration_TC)
                ws.cell(row = cnt+1, column =6, value = duration_frame)

                cnt += 1
        


        for row in ws.rows:
            for cell in row:
                if cell.value == 'Dissolve' :
                    cell.font = Font(color = "FF0000")
                cell.alignment = Alignment(
                    horizontal='center', vertical='center', wrap_text=True)

        output_file = os.path.join(os.path.expanduser('~'),'Desktop','edl_'+os.path.splitext(os.path.basename(edlfile))[0]+'.xlsx')

        wb.save(output_file)
        self.Done_e2x.show()



    def X2X_uploadXML(self):
        self.Done_x2x.hide()
        self.xmlpath.clear()
        global xmlfile
        xmlfile = QFileDialog.getOpenFileName(self, self.tr('Open File'), '/', 'All File(*);; xml File(*.xml)')[0]

        self.xmlpath.setText(xmlfile.split("/")[-1])

    def X2X_start(self) :
        wb = Workbook()
        ws = wb.active

        tree = qwe.parse(xmlfile)
        K = tree.find("sequence/media/video/track")
        A = K.findall('clipitem')

        file_name_li = []
        issue_li = []

        wb = Workbook()
        ws = wb.active 
                  #   1      2          3           4        5          6            7              8                   9               10           11          12       13
        col_names = ['', 'Reel ID', 'Clip name', 'TC in', 'TC out', 'Duration', 'Duration(f)','ORG_Duration', 'ORG_Duration(f)', 'Retime (%)', 'Resize (%)', 'note', 'note2']
        col_width = [6, 40, 40, 15, 15, 16, 17, 15, 16, 12, 12, 17, 17]

        for seq, name in enumerate(col_names):
            ws.cell(row=1, column=seq+1, value=name)
            ws.column_dimensions[get_column_letter(seq+1)].width = col_width[seq]
            ws.cell(row=1, column=seq+1).font = Font(bold=True)
            ws.cell(row=1, column=seq+1).fill = PatternFill(patternType='solid',fgColor=Color(indexed=44))

# ============================================SET FIN======================================================
        
        for num, clip in enumerate(A) :
            issue_li.clear()

            global TC_in
            global TC_out

            try :
                ws.cell(row=num+2, column=2, value=''.join(re.findall('[A-Z]\w+',clip.find('file').find('timecode').find('reel').findtext('name'))))
                file_name_li.append([clip.find('file').get('id'),
                    ''.join(re.findall('[A-Z]\w+',clip.find('file').find('timecode').find('reel').findtext('name'))),
                    clip.find('file').find('timecode').findtext('frame') if clip.find('file').find('timecode').findtext('frame') != None else 0 ])
                TC_in = frames_to_TC(int(clip.find('file').find('timecode').findtext('frame')) + int(clip.findtext('in'))) if clip.find('file').find('timecode').findtext('frame') != None else frames_to_TC(int(clip.findtext('in')))
                TC_out = frames_to_TC(int(clip.find('file').find('timecode').findtext('frame')) + int(clip.findtext('out'))) if clip.find('file').find('timecode').findtext('frame') != None else frames_to_TC(int(clip.findtext('out')))

                ws.cell(row = num+2, column= 4, value=TC_in)
                ws.cell(row=num+2, column=5, value=TC_out)

            except :
                for i in range(len(file_name_li)) :
                        if clip.find('file').get('id') == file_name_li[i][0] :
                            ws.cell(row=num+2, column = 2, value = file_name_li[i][1])
                            TC_in = frames_to_TC(int(file_name_li[i][2]) + int(clip.findtext('in')))
                            TC_out = frames_to_TC(int(file_name_li[i][2]) + int(clip.findtext('out')))
                            TC_duration = int(clip.findtext('end')) - int(clip.findtext('start'))

                            ws.cell(row = num+2, column= 4, value=TC_in)
                            ws.cell(row=num+2, column=5, value=TC_out)

            ws.cell(row=num+2, column=1, value=num+1) #Event
            ws.cell(row=num+2, column=3, value=clip.findtext('name')) #Clip Name

            TC_duration = int(clip.findtext('end')) - int(clip.findtext('start'))
            ws.cell(row = num+2, column=7, value = TC_duration if int(clip.findtext('start')) >= 0 and int(clip.findtext('end')) >= 0 else "Dissolve") #Duration frame
            ws.cell(row = num+2, column=6, value = frames_to_TC(TC_duration) if int(clip.findtext('start')) >= 0 and int(clip.findtext('end')) >=0 else "Dissolve")

            filter_li=clip.findall('filter')

            if len(filter_li) > 0 :
                for filter in filter_li :
                    effect_li = filter.find('effect').findall('parameter')

                    for effect in effect_li :
                        if effect.findtext('name') == 'Scale' and filter.find('effect').findtext('effectid') != 'GraphicAndType':
                            ws.cell(row=num+2, column=11, value=float(effect.findtext('value'))) # Resize
                            if effect.find('keyframe') is not None :
                                issue_li.append('* scale variable')
                        if effect.findtext('name') == 'speed' :
                            ws.cell(row=num+2, column=10, value=float(effect.findtext('value'))) #Retime
                            retime = float(effect.findtext('value'))
                
                        if effect.findtext('name') == 'reverse' :
                            if effect.findtext('value') == 'TRUE' :
                                issue_li.append('* reverse')
                        if effect.findtext('name') == 'graphdict' :
                            speed_key = effect.findall('keyframe')
                            for keyframe in speed_key :
                                # ws.cell(row=num+2, column=8, value =ws.cell(row=num+2, column=6).value) # EDIT duration 
                                # ws.cell(row=num+2, column=9, value=ws.cell(row=num+2, column=7).value) # EDIT duration (F)
                                new_du = ws.cell(row=num+2, column=7).value
                                if keyframe.findtext('when') == clip.findtext('in') : #TC_in 다시
                                    edit_du_in = float(keyframe.findtext('value'))
                                    if clip.find('file').findtext('name') != None :
                                        new_TC_in = int(clip.find('file').find('timecode').findtext('frame') if clip.find('file').find('timecode').findtext('frame') != None else 0) + edit_du_in
                                    else :
                                        for i in range(len(file_name_li)) :
                                            if clip.find('file').get('id') == file_name_li[i][0] :
                                                new_TC_in = int(file_name_li[i][2]) + edit_du_in

                            if new_du != "Dissolve" :
                                if retime < 100 :
                                    new_frame = math.ceil((retime/100) * new_du)
                                elif retime > 100 :
                                    new_frame = math.floor(retime/100 * new_du)
                                ws.cell(row = num + 2, column=8, value= frames_to_TC(new_frame))
                                ws.cell(row = num + 2, column=9, value= new_frame)
                                ws.cell(row=num+2, column=4, value=frames_to_TC(new_TC_in))
                                ws.cell(row=num+2, column=5, value=frames_to_TC(new_TC_in + int(new_frame)))
            issue = '\n'.join(s for s in issue_li)
            ws.cell(row=num+2, column=12, value=issue)

        for row in ws.rows:
            for cell in row:
                if cell.value == 'Dissolve' :
                    cell.font = Font(color = "FF0000")
                cell.alignment = Alignment(
                    horizontal='center', vertical='center', wrap_text=True)

        user_name = os.path.expanduser('~') + '/Desktop/'
        file_name = 'xml_list.xlsx'
        output = user_name + file_name
        uniq = 1

        while os.path.exists(output) :
            file_name = 'xml_list(%d).xlsx' % uniq
            output = user_name + file_name
            uniq += 1

        wb.save(output)
        self.Done_x2x.show()
 
    # def PC_uploadXL(self):

    # def PC_start(self):

    def BI_uploadXML(self):
        self.Done_BI.hide()
        self.xml_label_BI.clear()
        self.shot_name.clear()
        global xmlFile
        xmlFile = QFileDialog.getOpenFileName(self, self.tr('Open File'), '/', 'All File(*);; xml File(*.fcpxml)')[0]
        self.xml_label_BI.setText(xmlFile.split("/")[-1])
        self.xml_label_BI.show()

    def BI_uploadList(self):
        self.Done_BI.hide()
        self.list_label_BI.clear()
        global listFile
        listFile = QFileDialog.getOpenFileName(self, self.tr('Open File'), '/', 'All File(*);; text File(*.txt)')[0]
        self.list_label_BI.setText(listFile.split("/")[-1])
        self.list_label_BI.show()

    def BI_start(self):
        edited_lines = []
        shotName = self.shot_name.text()
        rename_wb = load_workbook(listFile, data_only=True)
        rename_ws = rename_wb.active
        new_lines = []
        for cell in rename_ws['A']:
            new_lines.append(cell.value)
        pattern = shotName + "_?\d{0,4}"
        print(len(new_lines))
        with open(xmlFile) as f:
            lines = f.readlines()
            i = 1
            for line in lines:
                if shotName in line and re.compile("ts\d{0,4}").search(line):
                    new_str = re.sub(pattern, new_lines[i], line)
                    edited_lines.append(new_str)
                    i += 1
                else:
                    edited_lines.append(line)

        user_name = os.path.expanduser('~')
        user_name = user_name + '/Desktop/burn_in_xml.xml'

        with open(user_name, 'w') as f:
            f.writelines(edited_lines)
        self.Done_BI.show()

    def RN_uploadFolder(self) :
        global folder
        folder = QFileDialog.getExistingDirectory(self, self.tr("Open Data files"), "/", QFileDialog.ShowDirsOnly)
        self.folder_label_RN.setText(folder.split("/")[-1])
        self.folder_label_RN.show()

    def RN_uploadList(self) :
        global rename_file
        rename_file = QFileDialog.getOpenFileName(self, self.tr('Open File'), '/', 'All File(*);; Excel File(*.xlsx)')[0]
        self.list_label_RN.setText(rename_file.split("/")[-1])
        self.list_label_RN.show()

    def RN_start(self) :
        self.Done_RN.hide()

        file_names = [k for k in os.listdir(folder) if not k.startswith('.')]
        wb = load_workbook(rename_file, data_only = True)
        ws = wb.active
        if len(file_names) != ws.max_row -1 :
            self.Done_RN.setText('FAILED.  (파일개수 : %d, 리네임리스트 수 : %d)'%(len(file_names),ws.max_row -1))
            self.Done_RN.show()
        else :
            for i in range(ws.max_row - 1) :
                old_name = os.path.join(folder, ws.cell(row = i+2, column=1).value)
                new_name = ws.cell(row = i+2, column=2).value if self.ext_path.text() == "" else ws.cell(column=2, row= i+2).value + "." + self.ext_path.text()
                new_name = os.path.join(folder,new_name)
                os.rename(old_name, new_name)

            self.Done_RN.show()


    def Size_uploadEXR(self) :
        global exrfolder
        exrfolder = QFileDialog.getExistingDirectory(self, self.tr("Open Data files"), "/", QFileDialog.ShowDirsOnly)
        self.exrpath.setText(exrfolder)


    def Size_start(self) :

        def print_resolution(folder_path, row):
            exr_files = [file for file in os.listdir(folder_path) if file.endswith(".exr")]

            exr_path = os.path.join(folder_path, exr_files[0])
    
            exr_file = OpenEXR.InputFile(exr_path)
            dw = exr_file.header()['dataWindow']
            width = dw.max.x - dw.min.x + 1
            height = dw.max.y - dw.min.y + 1
    
            folder_name = os.path.basename(folder_path)
            ws.cell(row=row, column=1, value=folder_name)
            ws.cell(row=row, column=2, value=f"{width} x {height}")
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            row += 1
            return row

        wb = Workbook()
        ws = wb.active
        ws['A1'] = '폴더 이름'
        ws['B1'] = '해상도'

# 첫번째 행 서식 적용
        for col in ws.iter_cols(min_row=1, max_row=1):
            for cell in col:
                cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
                cell.font = cell.font.copy(bold=True, color=Color(rgb='FFFFFF'))

                exrfolder_path = exrfolder
                row = 2
                for subfolder in os.listdir(exrfolder_path):
                    subfolder_path = os.path.join(exrfolder_path, subfolder)
                    if os.path.isdir(subfolder_path):
                        row = print_resolution(subfolder_path, row)
        
# 셀 간격 조정
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.row_dimensions[1].height = 20

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 엑셀 파일 저장
        output_file = os.path.join(os.path.expanduser('~'),'Desktop','exr_resolution.xlsx')
        wb.save(output_file)
        self.Done_size.show()


app = QApplication()
window = myWindow()
window.show()
app.exec()







