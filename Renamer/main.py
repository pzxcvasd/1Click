from PySide6.QtWidgets import *
from ui_1031 import Ui_MainWindow
from openpyxl import load_workbook
import os

class myWindow(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super(myWindow, self).__init__()
        self.setupUi(self)
        self.Done.hide()
        self.folder_select.clicked.connect(self.selectFolder)
        self.rename_select.clicked.connect(self.selectRenameFile)
        self.start.clicked.connect(self.startRename)

    def selectFolder(self) :
        global folder
        folder = QFileDialog.getExistingDirectory(self, self.tr("Open Data files"), "/", QFileDialog.ShowDirsOnly)
        self.folder_path.setText(folder)

    def selectRenameFile(self) :
        global rename_file
        rename_file = QFileDialog.getOpenFileName(self, self.tr('Open File'), '/', 'All File(*);; Excel File(*.xlsx)')[0]
        self.rename_path.setText(rename_file.split("/")[-1])

    def startRename(self) :
        file_names = [k for k in os.listdir(folder) if not k.startswith('.')]
        wb = load_workbook(rename_file, data_only = True)
        ws = wb.active
        if len(file_names) != ws.max_row -1 :
            self.Done.setStyleSheet("color : red")
            self.Done.setText('파일 개수(%d)와 리네임 개수(%d) 다름' % (len(file_names), ws.max_row -1))
            self.Done.show()
        else :
            for i in range(ws.max_row - 1) :
                old_name = os.path.join(folder, ws.cell(row = i+2, column=1).value)
                new_name = ws.cell(row = i+2, column=2).value if self.ext_path.text() == "" else ws.cell(column=2, row= i+2).value + "." + self.ext_path.text()
                new_name = os.path.join(folder,new_name)
                os.rename(old_name, new_name)

            self.Done.setStyleSheet('Color : green')
            self.Done.setText('Done')
            self.Done.show()

app = QApplication()
window = myWindow()
window.show()
app.exec()